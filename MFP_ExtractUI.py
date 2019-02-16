# UI Created by: PyQt5 UI code generator 5.5.1
#
# Adam T. Cuellar
# My 'Hello World' program for python :)

import myfitnesspal
import re
import plotly
import plotly.plotly as plot
import plotly.graph_objs as graph
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Fill, PatternFill
import easygui
from easygui import multpasswordbox
from easygui import enterbox
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import (QWidget, QProgressBar, QPushButton, QApplication)
from datetime import date
from dateutil.rrule import rrule, DAILY

# some global variables
startText = ""
finishText = ""
username = ""
password = ""

isNew = False

# graph things
def analyze(fileName):

    avgCals = 0
    avgProtein = 0
    avgFat = 0
    avgCarbs = 0
    avgSodium = 0
    filledLines = 0;

    # open existing excel file and retrieve first sheet
    book = load_workbook(fileName)
    macros_sheet = book['Macros']

    maxRow = macros_sheet.max_row

    # read info
    for i in range(5,maxRow):
        if macros_sheet.cell(i,2).value is not None:
            avgProtein += macros_sheet.cell(i,2).value
        if macros_sheet.cell(i,6).value is not None:
            avgCals += macros_sheet.cell(i,6).value
            filledLines += 1
        if macros_sheet.cell(i,4).value is not None:
            avgFat += macros_sheet.cell(i,4).value
        if macros_sheet.cell(i,3).value is not None:
            avgCarbs += macros_sheet.cell(i,3).value
        if macros_sheet.cell(i,7).value is not None:
            avgSodium += macros_sheet.cell(i,7).value

    avgCals /= filledLines
    avgProtein /= filledLines
    avgFat /= filledLines
    avgSodium /= filledLines
    avgCarbs /= filledLines

    plotly.offline.init_notebook_mode(connected=True)

    plotly.offline.plot({
        "data": [graph.Bar(x=["Total Cals", "Pro", "Carbs", "Fat", "Sodium"],
                               y=[avgCals, avgProtein, avgCarbs, avgFat, avgSodium],
                               text= [int(avgCals), int(avgProtein), int(avgCarbs), int(avgFat), int(avgSodium)],
                               textposition = 'auto')],
        "layout": graph.Layout(title="Averages")
    }, auto_open = True)


    return




# makes a new spreadsheet with cool format
def makeNewSpreadsheet(fileName, macros_sheet):

        macros_sheet['A1'].value = "Date"
        macros_sheet['A1'].font = Font(bold = True)
        macros_sheet['A1'].fill = PatternFill("solid", fgColor = "F08080")

        macros_sheet.merge_cells('B1:F1')
        macros_sheet['B1'].value = "Macros"
        macros_sheet['B1'].font = Font(bold = True)
        macros_sheet['B1'].fill = PatternFill("solid", fgColor = "F08080")

        macros_sheet.merge_cells('G1:H1')
        macros_sheet['G1'].value = "Electrolytes"
        macros_sheet['G1'].font = Font(bold = True)
        macros_sheet['G1'].fill = PatternFill("solid", fgColor = "F08080")

        macros_sheet.cell(2,1).value = ""
        macros_sheet.cell(2,1).font = Font(bold = True)
        macros_sheet.cell(2,1).fill = PatternFill("solid", fgColor = "D3D3D3")

        macros_sheet.cell(2,2).value = "Protein"
        macros_sheet.cell(2,2).font = Font(bold = True)
        macros_sheet.cell(2,2).fill = PatternFill("solid", fgColor = "D3D3D3")


        macros_sheet.cell(2,3).value = "Carbs"
        macros_sheet.cell(2,3).font = Font(bold = True)
        macros_sheet.cell(2,3).fill = PatternFill("solid", fgColor = "D3D3D3")


        macros_sheet.cell(2,4).value = "Fat"
        macros_sheet.cell(2,4).font = Font(bold = True)
        macros_sheet.cell(2,4).fill = PatternFill("solid", fgColor = "D3D3D3")


        macros_sheet.cell(2,5).value = "Fiber"
        macros_sheet.cell(2,5).font = Font(bold = True)
        macros_sheet.cell(2,5).fill = PatternFill("solid", fgColor = "D3D3D3")


        macros_sheet.cell(2,6).value = "Total Cals"
        macros_sheet.cell(2,6).font = Font(bold = True)
        macros_sheet.cell(2,6).fill = PatternFill("solid", fgColor = "D3D3D3")


        macros_sheet.cell(2,7).value = "SODIUM"
        macros_sheet.cell(2,7).font = Font(bold = True)
        macros_sheet.cell(2,7).fill = PatternFill("solid", fgColor = "D3D3D3")


        macros_sheet.cell(2,8).value = "POTASSIUM"
        macros_sheet.cell(2,8).font = Font(bold = True)
        macros_sheet.cell(2,8).fill = PatternFill("solid", fgColor = "D3D3D3")

        return

# create login GUI
def loginGUI():
        global username
        global password
        msg = "Enter MyFitnessPal Login Information"
        title = "Log In"
        fieldNames = ["Username", "Password"]
        fieldValues = []
        fieldValues = multpasswordbox(msg,title, fieldNames)

        # make sure each field gets an input
        while 1:
            if fieldValues is None:
                break
            errmsg = ""
            for i in range(len(fieldNames)):
              if fieldValues[i].strip() == "":
                errmsg = errmsg + ('"%s" is a required field.\n\n' % fieldNames[i])
            if errmsg == "": break # no problems found
            fieldValues = multpasswordbox(errmsg, title, fieldNames, fieldValues)

        username = fieldValues[0]
        password = fieldValues[1]


# adds information to excel sheet
def parseForExcel(dictionary, line, sh, date):

    # get each part of the dictionary
    carbs = dictionary.get("carbohydrates")
    pro = dictionary.get("protein")
    fat = dictionary.get("fat")
    cals = dictionary.get("calories")
    sodium = dictionary.get("sodium")
    sugar = dictionary.get("sugar")
    potassium = dictionary.get("potassium")


    # write info to respective column + row on sheet
    sh.cell(line,1).value = date
    sh.cell(line,2).value = pro
    sh.cell(line,3).value = carbs
    sh.cell(line,4).value = fat
    # row reserved for fiber, currently not working
    sh.cell(line,5).value = 0;
    sh.cell(line,6).value = cals
    sh.cell(line,7).value = sodium
    # riw reserved for potassium, currently not working
    sh.cell(line,8).value = 0

# extracts proper info from mfp
def extractInfo(StartDate,EndDate, fileName,self, createNew):

    completion = 5

    # if they're lame and don't already have a spreadsheet, make one for them
    # otherwise, proceed with chosen spreadsheet
    if(createNew is True):
        book = Workbook()
        macros_sheet = book.active
        macros_sheet.title = "Macros"
        makeNewSpreadsheet(fileName, macros_sheet)
        line = 3
    else:
        # open existing excel file and retrieve first sheet
        book = load_workbook(fileName)
        macros_sheet = book['Macros']

        # get last written in line number (skips 2 to leave extra space for in between weeks)
        line = macros_sheet.max_row + 2

    # open login UI
    loginGUI()

    client = None
    loginAttempts = 0;

    # make sure the user logs in correctly, if not give them 3 attempts
    while client is None:
        if loginAttempts is 3:
            easygui.msgbox("You've exceeded the maximum Log In attempts", "Warning")
            self.progress.setValue(0)
            return

        try:
            self.progress.setValue(completion)
            # assign account name to client and track for invalid credentials
            client = myfitnesspal.Client(username, password)
        except:
            loginAttempts += 1
            easygui.msgbox("Invalid Credentials. Please Try Again", "Warning")
            loginGUI()

    # split inputted dates
    StartDateArr = StartDate.split("/")
    EndDateArr = EndDate.split("/")

    # form date from input
    a = date(int(StartDateArr[2]),int(StartDateArr[0]),int(StartDateArr[1]))
    b = date(int(EndDateArr[2]),int(EndDateArr[0]),int(EndDateArr[1]))

    # calculate percentage increments for progress bar
    increment = 100/(b-a).days - 5/(b-a).days

    # iterate through each day
    for dt in rrule(DAILY, dtstart = a, until = b):
        year = int(dt.strftime("%Y"))
        month = int(dt.strftime("%m"))
        day = int(dt.strftime("%d"))
        dictionary = client.get_date(year, month, day).totals
        parseForExcel(dictionary,line, macros_sheet, dt.strftime("%-m/%-d/%Y"))
        line += 1
        completion += increment
        self.progress.setValue(completion)

    book.save(fileName)

    # show user we've completed the task
    easygui.msgbox("Task Completed!", "Success")
    # reset progress bar
    self.progress.setValue(0)
    return

# GUI setup
class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(612, 409)
        self.centralWidget = QtWidgets.QWidget(MainWindow)
        self.centralWidget.setObjectName("centralWidget")
        self.calendarWidget = QtWidgets.QCalendarWidget(self.centralWidget)
        self.calendarWidget.setGeometry(QtCore.QRect(20, 20, 270, 200))
        self.progress = QProgressBar(self.centralWidget)
        self.progress.setGeometry(QtCore.QRect(18, 255, 270, 20))
        self.calendarWidget.setObjectName("calendarWidget")
        self.pushButton = QtWidgets.QPushButton(self.centralWidget)
        self.pushButton.setGeometry(QtCore.QRect(400, 250, 131, 27))
        self.pushButton.setObjectName("pushButton")
        self.pushButton_2 = QtWidgets.QPushButton(self.centralWidget)
        self.pushButton_2.setGeometry(QtCore.QRect(400, 290, 131, 27))
        self.pushButton_2.setObjectName("pushButton_2")
        self.pushButton_3 = QtWidgets.QPushButton(self.centralWidget)
        self.pushButton_3.setGeometry(QtCore.QRect(400, 330, 131, 27))
        self.pushButton_3.setObjectName("pushButton_3")
        self.lineEdit = QtWidgets.QLineEdit(self.centralWidget)
        self.lineEdit.setGeometry(QtCore.QRect(470, 40, 113, 27))
        self.lineEdit.setObjectName("lineEdit")
        self.lineEdit_2 = QtWidgets.QLineEdit(self.centralWidget)
        self.lineEdit_2.setGeometry(QtCore.QRect(470, 80, 113, 27))
        self.lineEdit_2.setObjectName("lineEdit_2")
        self.label = QtWidgets.QLabel(self.centralWidget)
        self.label.setGeometry(QtCore.QRect(300, 40, 171, 20))
        self.label.setObjectName("label")
        self.label_2 = QtWidgets.QLabel(self.centralWidget)
        self.label_2.setGeometry(QtCore.QRect(320, 80, 151, 20))
        self.label_2.setObjectName("label_2")
        self.label_3 = QtWidgets.QLabel(self.centralWidget)
        self.label_3.setObjectName("label_3")
        self.label_3.setGeometry(QtCore.QRect(18, 235, 170, 20))
        MainWindow.setCentralWidget(self.centralWidget)
        self.mainToolBar = QtWidgets.QToolBar(MainWindow)
        self.mainToolBar.setObjectName("mainToolBar")
        MainWindow.addToolBar(QtCore.Qt.TopToolBarArea, self.mainToolBar)
        self.statusBar = QtWidgets.QStatusBar(MainWindow)
        self.statusBar.setObjectName("statusBar")
        MainWindow.setStatusBar(self.statusBar)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MFP Extractor"))
        self.pushButton.setText(_translate("MainWindow", "Edit Spreadsheet"))
        self.pushButton_2.setText(_translate("MainWindow", "Create Spreadsheet"))
        self.pushButton_3.setText(_translate("MainWindow", "Analyze Spreadsheet"))
        self.label.setText(_translate("MainWindow", "Starting Date (MM/DD/YYYY): "))
        self.label_2.setText(_translate("MainWindow", "End Date (MM/DD/YYYY): "))
        self.label_3.setText(_translate("MainWindow", "Progress:"))
        self.pushButton.clicked.connect(self.clickedOpenSpreadsheet)
        self.pushButton_2.clicked.connect(self.clickedNewSpreadsheet)
        self.pushButton_3.clicked.connect(self.clickedAnalyzeSpreadsheet)


    # read and redirect inputted dates
    def clickedOpenSpreadsheet(self):
        global startText
        global finishText
        global isNew
        startText = self.lineEdit.text()
        finishText = self.lineEdit_2.text()
        x = re.search("[0-9]{2}/[0-9]{2}/[0-9]{4}",startText);
        y = re.search("[0-9]{2}/[0-9]{2}/[0-9]{4}",finishText);

        # warn user if date format is incorrect, else open find file dialog
        if x is None or y is None:
            easygui.msgbox("Incorrect Date Format", "Warning")
        else:
            fileName = easygui.fileopenbox()
            extractInfo(startText,finishText,fileName,self, False)

        return

    # same as above except make a new spreadsheet
    def clickedNewSpreadsheet(self):
        global startText
        global finishText
        global isNew
        startText = self.lineEdit.text()
        finishText = self.lineEdit_2.text()
        x = re.search("[0-9]{2}/[0-9]{2}/[0-9]{4}",startText);
        y = re.search("[0-9]{2}/[0-9]{2}/[0-9]{4}",finishText);

        # warn user if date format is incorrect, else open find file dialog
        if x is None or y is None:
            easygui.msgbox("Incorrect Date Format", "Warning")
        else:
            fileName = enterbox("Enter a name for the new spreadsheet:") + ".xlsx"
            extractInfo(startText,finishText,fileName,self, True)

        return

    # open a spreadsheet
    def clickedAnalyzeSpreadsheet(self):
        fileName = easygui.fileopenbox()
        analyze(fileName)



if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())


